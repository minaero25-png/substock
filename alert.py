"""
Sub Stock Expiry Alert — Telegram Bot
อ่าน data.xlsx → หายาใกล้หมดอายุ → ส่งแจ้งเตือน Telegram
"""
import os
import sys
from datetime import datetime, timedelta
import urllib.request
import urllib.parse
import json

# --- CONFIG ---
EXCEL_FILE = "data.xlsx"
SHEET_NAME = "Sub stock Data"
ALERT_DAYS = 180  # แจ้งเตือนยาที่จะหมดอายุภายใน 180 วัน (6 เดือน)
URGENT_DAYS = 30  # ยาที่หมดอายุภายใน 30 วัน = ด่วนมาก

# --- Read secrets from environment ---
BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")

def send_telegram(text):
    """ส่งข้อความไป Telegram"""
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    data = urllib.parse.urlencode({
        "chat_id": CHAT_ID,
        "text": text,
        "parse_mode": "HTML"
    }).encode()
    req = urllib.request.Request(url, data=data)
    try:
        with urllib.request.urlopen(req) as res:
            return res.status == 200
    except Exception as e:
        print(f"Telegram send failed: {e}")
        return False

def main():
    if not BOT_TOKEN or not CHAT_ID:
        print("ERROR: Missing TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID")
        sys.exit(1)

    # --- Read Excel ---
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl -q")
        import openpyxl

    if not os.path.exists(EXCEL_FILE):
        send_telegram("⚠️ <b>Sub Stock Alert</b>\n\nไม่พบไฟล์ data.xlsx ใน repo กรุณาอัปโหลดไฟล์ข้อมูล")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]

    # --- Parse header ---
    headers = [cell.value for cell in ws[1]]
    def col(name):
        try: return headers.index(name)
        except: return -1

    ci_name = col("รายการทั้งหมด")
    ci_lot  = col("Lot number")
    ci_exp  = col("วันหมดอายุ")
    ci_qty  = col("จำนวน")
    ci_type = col("ประเภท")
    ci_act  = col("สถานะ")

    if ci_name < 0 or ci_exp < 0:
        send_telegram("⚠️ <b>Sub Stock Alert</b>\n\nไม่พบคอลัมน์ 'รายการทั้งหมด' หรือ 'วันหมดอายุ' ใน Excel")
        sys.exit(1)

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    alert_date = today + timedelta(days=ALERT_DAYS)
    urgent_date = today + timedelta(days=URGENT_DAYS)

    expired = []    # หมดอายุแล้ว
    urgent = []     # ด่วน (≤30 วัน)
    warning = []    # เตือน (31-180 วัน)

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[ci_name] if ci_name >= 0 else None
        if not name:
            continue

        exp = row[ci_exp] if ci_exp >= 0 else None
        qty = row[ci_qty] if ci_qty >= 0 else 0
        lot = row[ci_lot] if ci_lot >= 0 else ""
        typ = row[ci_type] if ci_type >= 0 else ""
        act = row[ci_act] if ci_act >= 0 else None

        # ข้ามรายการที่จัดการแล้ว
        if act:
            continue

        # ข้าม stock = 0
        qty = qty or 0
        if qty == 0:
            continue

        if not isinstance(exp, datetime):
            continue

        days_left = (exp - today).days

        item = {
            "name": str(name).strip(),
            "lot": str(lot).strip(),
            "exp": exp.strftime("%Y-%m-%d"),
            "days": days_left,
            "qty": int(qty),
            "type": str(typ).strip()
        }

        if days_left < 0:
            expired.append(item)
        elif days_left <= URGENT_DAYS:
            urgent.append(item)
        elif days_left <= ALERT_DAYS:
            warning.append(item)

    # --- Build message ---
    total_alerts = len(expired) + len(urgent) + len(warning)

    if total_alerts == 0:
        msg = "💊 <b>Sub Stock Daily Report</b>\n"
        msg += f"📅 {today.strftime('%d/%m/%Y')}\n\n"
        msg += "✅ ไม่มียาที่ต้องเฝ้าระวังในขณะนี้\n"
        msg += f"(ตรวจสอบยาที่หมดอายุภายใน {ALERT_DAYS} วัน)"
        send_telegram(msg)
        print("No alerts. Message sent.")
        return

    msg = "💊 <b>Sub Stock — แจ้งเตือนยาใกล้หมดอายุ</b>\n"
    msg += f"📅 {today.strftime('%d/%m/%Y')}\n"
    msg += f"⏰ ตรวจสอบยาที่หมดอายุภายใน {ALERT_DAYS} วัน\n"
    msg += "━━━━━━━━━━━━━━━━━━\n"

    if expired:
        msg += f"\n🔴 <b>หมดอายุแล้ว ({len(expired)} รายการ)</b>\n"
        for i in expired:
            msg += f"  • {i['name']}\n"
            msg += f"    Lot: {i['lot']} | Exp: {i['exp']} ({abs(i['days'])} วันที่แล้ว)\n"
            msg += f"    คงเหลือ: {i['qty']} | {i['type']}\n"

    if urgent:
        msg += f"\n🟠 <b>ด่วน! หมดอายุใน ≤{URGENT_DAYS} วัน ({len(urgent)} รายการ)</b>\n"
        for i in urgent:
            msg += f"  • {i['name']}\n"
            msg += f"    Lot: {i['lot']} | Exp: {i['exp']} (อีก {i['days']} วัน)\n"
            msg += f"    คงเหลือ: {i['qty']} | {i['type']}\n"

    if warning:
        msg += f"\n🟡 <b>เตือน: หมดอายุใน {URGENT_DAYS+1}-{ALERT_DAYS} วัน ({len(warning)} รายการ)</b>\n"
        for i in warning:
            msg += f"  • {i['name']}\n"
            msg += f"    Lot: {i['lot']} | Exp: {i['exp']} (อีก {i['days']} วัน)\n"

    msg += "\n━━━━━━━━━━━━━━━━━━\n"
    msg += f"📊 สรุป: 🔴{len(expired)} 🟠{len(urgent)} 🟡{len(warning)} รวม {total_alerts} รายการ\n"
    msg += "🔗 Dashboard: https://minaero25-png.github.io/substock/"

    # Telegram message limit is 4096 chars
    if len(msg) > 4000:
        # Send summary only
        msg = "💊 <b>Sub Stock — แจ้งเตือนยาใกล้หมดอายุ</b>\n"
        msg += f"📅 {today.strftime('%d/%m/%Y')}\n"
        msg += "━━━━━━━━━━━━━━━━━━\n"
        msg += f"🔴 หมดอายุแล้ว: {len(expired)} รายการ\n"
        msg += f"🟠 ด่วน (≤{URGENT_DAYS} วัน): {len(urgent)} รายการ\n"
        msg += f"🟡 เตือน ({URGENT_DAYS+1}-{ALERT_DAYS} วัน): {len(warning)} รายการ\n"
        msg += f"\n📊 รวม {total_alerts} รายการที่ต้องเฝ้าระวัง\n"
        msg += "⚠️ รายการเยอะเกินจะแสดงทั้งหมด กรุณาดูรายละเอียดที่ Dashboard\n"
        msg += "🔗 https://minaero25-png.github.io/substock/"

    success = send_telegram(msg)
    print(f"Sent: {total_alerts} alerts, Telegram: {'OK' if success else 'FAILED'}")

if __name__ == "__main__":
    main()
